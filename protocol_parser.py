"""
Parser for T4V Public Data Protocol V202 xlsx files.

Reads product data from the structured xlsx sheets (Product, Material, Care,
Brand, Supply chain) and returns unified product records ready for DB storage.
"""
import json
import re

from openpyxl import load_workbook


def _read_sheet(wb, sheet_name):
    """Read a sheet into a list of dicts using the first row as headers."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if not any(row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _extract_color_name_and_code(color_str):
    """Extract color name and code from format like 'Off-black (9073)'.

    Returns (color_name, color_code) tuple.
    """
    if not color_str:
        return None, None
    match = re.match(r"(.+?)\s*\((\d+)\)", str(color_str))
    if match:
        return match.group(1).strip(), match.group(2)
    return str(color_str).strip(), None


def _make_product_key(article_number, color_code):
    """Build the article+colorcode key used in Material/Care/Brand/Supply chain sheets."""
    if not article_number or not color_code:
        return None
    return f"{article_number}{color_code}"


def parse_protocol_xlsx(filepath):
    """Parse a T4V protocol xlsx file and return a list of product dicts.

    Each dict has keys matching the products_v2 table columns:
    gtin, article_number, product_name, description, category, size, color,
    materials (JSON string), care_text, brand, country_of_origin.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # ── Read Product sheet ───────────────────────────────────────────
    products_raw = _read_sheet(wb, "Product")

    # ── Read Material sheet → group by product key ───────────────────
    materials_raw = _read_sheet(wb, "Material")
    materials_by_key = {}
    for row in materials_raw:
        key = str(row.get("Product Identification Value (GTIN, SKU ID value, Style ID value)", "")).strip()
        if not key:
            continue
        mat_entry = {
            "name": str(row.get("material Content Name", "") or "").strip(),
            "percentage": _parse_percentage(row.get("Content Value (material Composition)")),
            "component": str(row.get("Component", "") or "").strip(),
        }
        materials_by_key.setdefault(key, []).append(mat_entry)

    # ── Read Care sheet → index by product key ───────────────────────
    care_raw = _read_sheet(wb, "Care")
    care_by_key = {}
    for row in care_raw:
        key = str(row.get("Product Identification Value (GTIN, SKU ID value, Style ID value)", "")).strip()
        if not key:
            continue
        care_by_key[key] = str(row.get("Care Text", "") or "").strip()

    # ── Read Brand sheet → index by product key ──────────────────────
    brand_raw = _read_sheet(wb, "Brand")
    brand_by_key = {}
    for row in brand_raw:
        key = str(row.get("Product Identification Value (GTIN, SKU ID value, Style ID value)", "")).strip()
        if not key:
            continue
        brand_by_key[key] = str(row.get("Brand", "") or "").strip()

    # ── Read Supply chain sheet → index by product key ───────────────
    supply_raw = _read_sheet(wb, "Supply chain")
    origin_by_key = {}
    for row in supply_raw:
        key = str(row.get("Product Identification Value (GTIN, SKU ID value, Style ID value)", "")).strip()
        if not key:
            continue
        origin_by_key[key] = str(row.get("Country of Origin - Confection", "") or "").strip()

    wb.close()

    # ── Build unified product records ────────────────────────────────
    results = []
    for prod in products_raw:
        gtin = str(prod.get("Product Identification Value (GTIN, SKU ID value, Style ID value)", "") or "").strip()
        article_number = str(prod.get("Article Number", "") or "").strip()
        product_name = str(prod.get("Product Name", "") or "").strip()
        description = str(prod.get("Consumer-Facing Description (Detailed)", "") or "").strip()
        category = str(prod.get("Category", "") or "").strip()
        size = str(prod.get("Size", "") or "").strip()
        color_raw = str(prod.get("Color (Brand)", "") or "").strip()

        if not gtin:
            continue

        color_name, color_code = _extract_color_name_and_code(color_raw)
        product_key = _make_product_key(article_number, color_code)

        materials = materials_by_key.get(product_key, []) if product_key else []
        care_text = care_by_key.get(product_key, "") if product_key else ""
        brand = brand_by_key.get(product_key, "") if product_key else ""
        country = origin_by_key.get(product_key, "") if product_key else ""

        results.append({
            "gtin": gtin,
            "article_number": article_number,
            "product_name": product_name,
            "description": description,
            "category": category,
            "size": size,
            "color": color_name or color_raw,
            "materials": json.dumps(materials) if materials else "[]",
            "care_text": care_text,
            "brand": brand,
            "country_of_origin": country,
        })

    return results


def _parse_percentage(value):
    """Parse a percentage value that may be a decimal (0.99) or integer (99)."""
    if value is None:
        return 0.0
    try:
        val = float(value)
        # If > 1, assume it's already a percentage (e.g. 99 → 0.99)
        if val > 1:
            return val / 100
        return val
    except (ValueError, TypeError):
        return 0.0
